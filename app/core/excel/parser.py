from decimal import Decimal, InvalidOperation
from datetime import datetime, date
import openpyxl

from .mapping import ColumnaExcel


def parsear_excel(archivo, mapeo: list[ColumnaExcel]):
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active

    fila_datos_inicio = max(col.fila_encabezado for col in mapeo) + 1

    data = []
    errores = []

    for fila_num in range(fila_datos_inicio, ws.max_row + 1):
        if all(ws[col.celda_en_fila(fila_num)].value is None for col in mapeo):
            break

        fila_data = {}
        for col in mapeo:
            celda_ref = col.celda_en_fila(fila_num)
            valor_raw = ws[celda_ref].value

            if valor_raw is None:
                tiene_error = col.requerido
                fila_data[col.campo] = {
                    'celda': celda_ref,
                    'valor': None,
                    'tipo': col.tipo,
                    'error': tiene_error,
                }
                if tiene_error:
                    errores.append({
                        'celda': celda_ref,
                        'campo': col.campo,
                        'valor': None,
                        'descripcion': f"El campo '{col.campo}' es requerido y está vacío.",
                    })
                continue

            valor_cast, error_msg = _castear(valor_raw, col.tipo)
            tiene_error = error_msg is not None

            fila_data[col.campo] = {
                'celda': celda_ref,
                'valor': valor_raw if tiene_error else valor_cast,
                'tipo': col.tipo,
                'error': tiene_error,
            }

            if tiene_error:
                errores.append({
                    'celda': celda_ref,
                    'campo': col.campo,
                    'valor': valor_raw,
                    'descripcion': error_msg,
                })

        data.append(fila_data)

    return data, errores


def _castear(valor, tipo: str):
    """Retorna (valor_casteado, error_msg). error_msg es None si no hubo error."""
    try:
        if tipo == 'string':
            return str(valor).strip(), None

        if tipo == 'integer':
            if isinstance(valor, float):
                if not valor.is_integer():
                    raise ValueError()
                return int(valor), None
            return int(valor), None

        if tipo == 'decimal':
            return str(Decimal(str(valor))), None

        if tipo == 'date':
            if isinstance(valor, (datetime, date)):
                v = valor.date() if isinstance(valor, datetime) else valor
                return v.isoformat(), None
            raise ValueError()

        return str(valor), None

    except (ValueError, TypeError, InvalidOperation):
        return None, f"Se esperaba '{tipo}' pero se recibió '{valor}'."
